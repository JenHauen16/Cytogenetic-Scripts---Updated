#use drop list updated on fri and sat 

import os
import re
import openpyxl
import datetime
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta, MO
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Border, Side
from openpyxl.worksheet import page
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
wsname = "GSL Cases for " + str(date.today())
ws = wb.create_sheet(index=0, title=wsname)
#wb.create_sheet(index = 1, title="Completed Cases")
#ws2 = wb["ML"]

ws['B1'] = "Tray"
ws['C1'] = "ID"
ws['D1'] = "Patient Name"
ws['E1'] = "Specimen Type"
ws['F1'] = 'Cultures'

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 25
#ws.column_dimensions['F'].width = 11
#ws.column_dimensions['G'].width = 11
#ws.row_dimensions[3].height = 65


#file = "ChromosomeCases Week of_2019-08-05.xlsx"
#os.chdir('Y:\\documents\\PATH-LABS\\CYTOGENETICS\\CYTOGEN\\Daily Log\\New DropLists')
#today = date.today()
#last_monday = today + relativedelta(weekday=MO(-1))
#file = "ChromosomeCases Week of_" + str(last_monday) + ".xlsx"
prevmon = date.today()- timedelta(days=7)
os.chdir('S:\\shares\\Pathology\\documents\\PATH-LABS\\CYTOGENETICS\\CYTOGEN\\Daily Log\\New DropLists')
file = "ChromosomeCases Week of_" + str(date.today()- timedelta(days=7)) + ".xlsx"
wb2=openpyxl.load_workbook(file)
print(wb2.sheetnames)

#print(wb2.sheetnames) str(date.today().strftime("%Y-%m-%d")-timedelta(days=1))
#ws2=wb2[str(date.today()- timedelta(days=1))]
ws2=wb2[str(prevmon)]

#ws2max = int(ws2.max_row)
#print(ws2max)
#file = 'Daily Log' + date.today().strftime("%m.%d.%y") + '.xlsx'

for index, row in enumerate(ws2.iter_rows()):
    #print(str(index))
    #print(row[6].value)
    #continue;
    #print(row[4].value)
    #print(row[5].value)
    new_line_num = ws.max_row + 1
    try:
        if "S" in row[0].value or "CG" in row[0].value or "FL" in row[0].value:
            #print(ws2.cell(row=index+2, column=5).value)
            #print(type(ws2.cell(row=index+2, column=5).value))
            #print(row[5].value)
            #print(ws2.cell(row=index+2, column=5).value.date());
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=5).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=5).value.date():
                #print("hi")
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=6, row=new_line_num, value=row[4].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=6).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=6).value.date():
                #print("no")
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=7, row=new_line_num, value=row[5].value)
                print(str(row[0].value + " " + row[5].value))
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=7).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=7).value.date():
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=8, row=new_line_num, value=row[6].value)
                print(row[0].value + " " + row[6].value)
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=8).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=8).value.date():    
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=9, row=new_line_num, value=row[7].value)
                print(row[0].value + " " + row[7].value)
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=9).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=9).value.date():
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=10, row=new_line_num, value=row[8].value)
                print(row[0].value + " " + row[8].value)
            if datetime.date.today()-timedelta(days=2) == ws2.cell(row=index+2, column=10).value.date() or datetime.date.today()-timedelta(days=3) == ws2.cell(row=index+2, column=10).value.date():
                ws.cell(column=3, row=new_line_num, value=row[0].value)
                ws.cell(column=4, row=new_line_num, value=row[1].value)
                ws.cell(column=5, row=new_line_num, value=row[2].value)
                ws.cell(column=11, row=new_line_num, value=row[9].value)
                print(row[0].value + " " + row[6].value)      
    except (TypeError, AttributeError):
        continue

bold12Font = Font(size=12, bold=True)
border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

row1 = ws["A1:G1"]
for c in row1:
    for a in c:
        a.font = bold12Font

rows = ws["A1:K36"]
for row in rows:
    for cell in row:
        cell.border = border
        #cell.font = bold12Font
        #cell.alignment = Alignment(horizontal="center")
        #cell.font = "Tahoma,Bold"
row2 = ws["A2:A31"]
for row in row2:
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

#row3 = ws["E2:E30"]
#for row in row3:
 #   for cell in row:
  #      cell2 = str(cell.value)
   #     cell2.replace("Bone marrow", "BM")
    #    cell2.replace("Peripheral blood", "PB")
        
ws['A2'] = 1
ws['A3'] = 2
ws['A4'] = 3
ws['A5'] = 4
ws['A6'] = 5
ws['A7'] = 6
ws['A8'] = 7
ws['A9'] = 8
ws['A10'] = 9
ws['A11'] = 10
ws['A12'] = 11
ws['A13'] = 12
ws['A14'] = 13
ws['A15'] = 14
ws['A16'] = 15
ws['A17'] = 16
ws['A18'] = 17
ws['A19'] = 18
ws['A20'] = 19
ws['A21'] = 20
ws['A22'] = 21
ws['A23'] = 22
ws['A24'] = 23
ws['A25'] = 24
ws['A26'] = 25
ws['A27'] = 26
ws['A28'] = 27
ws['A29'] = 28
ws['A30'] = 29
ws['A31'] = 30

ws.oddHeader.left.text = 'GSL List_' + str(date.today()) 
ws.oddHeader.left.size = 12
ws.oddHeader.left.font = "Tahoma,Bold"
#ws.oddHeader.center.color = "CC3366"

ws.oddHeader.right.text = "Tech Loading: ________________ \r Tech Checking: ________________"
ws.oddHeader.right.size = 12
ws.oddHeader.right.font = "Tahoma,Bold"

#ws.oddHeader.center.text = "Date: ________________"
#ws.oddHeader.center.size = 12
#ws.oddHeader.center.font = "Tahoma,Bold"

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

wb.save('GSL Case List' + '_' + str(date.today()) + '.xlsx')
wb.close()

#wb.remove(wb['Sheet'])
#ws.delete_rows(3,4)

#for row in ws.rows:
 #   cell = row[:8][1]
  #  cell1 = row[:8][2]
   # cell2 = row[:8][3]
   # cell3 = row[:8][4]
    #cell4 = row[8:][0]
#    cell.alignment = Alignment(horizontal="center")
 #   cell1.alignment = Alignment(horizontal="center")
  #  cell2.alignment = Alignment(horizontal="center")
   # cell3.alignment = Alignment(horizontal="center")
   # cell4.alignment = Alignment(horizontal="center")'''

#Font code that works!
#bold14Font = Font(size=14, bold=True)
#ws['A1'].font = bold14Font
#ws['A1'].fill = PatternFill(fgColor="CC33FF", fill_type="solid")
#ws['A2'].font = bold14Font
#ws['B2'].font = bold14Font
#ws['C2'].font = bold14Font



