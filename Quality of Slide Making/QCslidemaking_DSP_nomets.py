#use updated drop list after tech records date, tech, & quality. use same day

import os
import re
import openpyxl
import datetime
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta, MO
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Border, Side
from openpyxl.worksheet import page
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
wsn = wb.create_sheet(index=0, title="No Mets")
ws = wb.create_sheet(index=1, title="Poor")
wsf = wb.create_sheet(index=2, title="Fair")
wsg = wb.create_sheet(index=3, title="Good")

#wsname = "Poor"
#wsname = "Slide_QC_" + str(date.today()+ timedelta(days=1))
#wb.create_sheet(index = 1, title="Completed Cases")
#ws2 = wb["ML"]

wsn['A1'] = "Number"
wsn['B1'] = "Path_ID"
wsn['C1'] = "Patient_Name"
wsn['D1'] = "Specimen Type"
wsn['E1'] = 'Cultures'

wsn.column_dimensions['A'].width = 9
wsn.column_dimensions['B'].width = 12
wsn.column_dimensions['C'].width = 20
wsn.column_dimensions['D'].width = 27
wsn.column_dimensions['E'].width = 12
wsn.column_dimensions['F'].width = 11
wsn.column_dimensions['G'].width = 11
#ws.row_dimensions[3].height = 65

ws['A1'] = "Number"
ws['B1'] = "Path_ID"
ws['C1'] = "Patient_Name"
ws['D1'] = "Specimen Type"
ws['E1'] = 'Cultures'

ws.column_dimensions['A'].width = 9
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 27
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 11
ws.column_dimensions['G'].width = 11
#ws.row_dimensions[3].height = 65

wsf['A1'] = "Number"
wsf['B1'] = "Path_ID"
wsf['C1'] = "Patient_Name"
wsf['D1'] = "Specimen Type"
wsf['E1'] = 'Cultures'

wsf.column_dimensions['A'].width = 9
wsf.column_dimensions['B'].width = 12
wsf.column_dimensions['C'].width = 20
wsf.column_dimensions['D'].width = 27
wsf.column_dimensions['E'].width = 12
wsf.column_dimensions['F'].width = 11
wsf.column_dimensions['G'].width = 11

wsg['A1'] = "Number"
wsg['B1'] = "Path_ID"
wsg['C1'] = "Patient_Name"
wsg['D1'] = "Specimen Type"
wsg['E1'] = 'Cultures'

wsg.column_dimensions['A'].width = 9
wsg.column_dimensions['B'].width = 12
wsg.column_dimensions['C'].width = 20
wsg.column_dimensions['D'].width = 27
wsg.column_dimensions['E'].width = 12
wsg.column_dimensions['F'].width = 11
wsg.column_dimensions['G'].width = 11

os.chdir('S:\\shares\\Pathology\\documents\\PATH-LABS\\CYTOGENETICS\\CYTOGEN\\Daily Log\\New DropLists\\February 2020')
#file = "ChromosomeCases Week of_" + str(date.today()- timedelta(days=1)) + ".xlsx"
#file = "ChromosomeCases Week of_2019-12-30.xlsx"
today = date.today()
last_monday = today + relativedelta(weekday=MO(-2))
file = "ChromosomeCases Week of_" + str(last_monday) + ".xlsx"
wb2=openpyxl.load_workbook(file)
print(wb2.sheetnames)
ws2=wb2["Completed Cases"]
#ws2=wb2[str(last_monday)]

#print(wb2.sheetnames) str(date.today().strftime("%Y-%m-%d")-timedelta(days=1))
#ws2max = int(ws2.max_row)
#print(ws2max)

for index, row in enumerate(ws2.iter_rows()):
    #print(str(index))
    #print(row[6].value)
    new_line_num = ws.max_row + 1
    nlwsg = wsg.max_row + 1
    nlwsf = wsf.max_row + 1
    nlwsn = wsn.max_row + 1
    try:
        if  "DSP" in row[4].value or "DSP" in row[5].value or "DSP" in row[6].value:
            #print(ws2.cell(row=index+2, column=5).value)
            #print(type(ws2.cell(row=index+2, column=5).value))
            #print(row[5].value)
            #print(ws2.cell(row=index+2, column=5).value.date());
            if "good" == ws2.cell(row=index+4, column=5).value:
                #print("hi")
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=5, row=nlwsg, value=row[4].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "fair" == ws2.cell(row=index+4, column=5).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=5, row=nlwsf, value=row[4].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=5).value:
                #print("hi")
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=5, row=new_line_num, value=row[4].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "no mets" == ws2.cell(row=index+4, column=5).value:
                #print("hi")
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=5, row=nlwsn, value=row[4].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
    
            if "good" == ws2.cell(row=index+4, column=6).value:
                #print("no")
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=6, row=nlwsg, value=row[5].value)
                print(str(row[0].value + " " + row[5].value))
            if "fair" == ws2.cell(row=index+4, column=6).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=6, row=nlwsf, value=row[5].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=6).value:
                #print("no")
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=6, row=new_line_num, value=row[5].value)
                print(str(row[0].value + " " + row[5].value))
            if "no mets" == ws2.cell(row=index+4, column=6).value:
                #print("no")
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=6, row=nlwsn, value=row[5].value)
                print(str(row[0].value + " " + row[5].value))
                
            if "good" == ws2.cell(row=index+4, column=7).value:
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=7, row=nlwsg, value=row[6].value)
                print(row[0].value + " " + row[6].value)
            if "fair" == ws2.cell(row=index+4, column=7).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=7, row=nlwsf, value=row[6].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=7).value:
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=7, row=new_line_num, value=row[6].value)
                print(row[0].value + " " + row[6].value)
            if "no mets" == ws2.cell(row=index+4, column=7).value:
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=7, row=nlwsn, value=row[6].value)
                print(row[0].value + " " + row[6].value)
                
            if "good" == ws2.cell(row=index+4, column=8).value:    
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=8, row=nlwsg, value=row[7].value)
                print(row[0].value + " " + row[7].value)
            if "fair" == ws2.cell(row=index+4, column=8).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=8, row=nlwsf, value=row[7].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=8).value:    
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=8, row=new_line_num, value=row[7].value)
                print(row[0].value + " " + row[7].value)
            if "no mets" == ws2.cell(row=index+4, column=8).value:    
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=8, row=nlwsn, value=row[7].value)
                print(row[0].value + " " + row[7].value)
                
            if "good" == ws2.cell(row=index+4, column=9).value:
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=9, row=nlwsg, value=row[8].value)
                print(row[0].value + " " + row[8].value)
            if "fair" == ws2.cell(row=index+4, column=9).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=9, row=nlwsf, value=row[8].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=9).value:
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=9, row=new_line_num, value=row[8].value)
                print(row[0].value + " " + row[8].value)
            if "no mets" == ws2.cell(row=index+4, column=9).value:
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=9, row=nlwsn, value=row[8].value)
                print(row[0].value + " " + row[8].value)

            if "good" == ws2.cell(row=index+4, column=10).value:
                wsg.cell(column=2, row=nlwsg, value=row[0].value)
                wsg.cell(column=3, row=nlwsg, value=row[1].value)
                wsg.cell(column=4, row=nlwsg, value=row[2].value)
                wsg.cell(column=10, row=nlwsg, value=row[9].value)
                print(row[0].value + " " + row[6].value)
            if "fair" == ws2.cell(row=index+4, column=10).value:
                #print("hi")
                wsf.cell(column=2, row=nlwsf, value=row[0].value)
                wsf.cell(column=3, row=nlwsf, value=row[1].value)
                wsf.cell(column=4, row=nlwsf, value=row[2].value)
                wsf.cell(column=10, row=nlwsf, value=row[9].value)
                print(str(row[0].value) + " " + row[1].value + " " + row[2].value + " " + row[4].value)
            if "poor" == ws2.cell(row=index+4, column=10).value:
                ws.cell(column=2, row=new_line_num, value=row[0].value)
                ws.cell(column=3, row=new_line_num, value=row[1].value)
                ws.cell(column=4, row=new_line_num, value=row[2].value)
                ws.cell(column=10, row=new_line_num, value=row[9].value)
                print(row[0].value + " " + row[6].value)
            if "no mets" == ws2.cell(row=index+4, column=10).value:
                wsn.cell(column=2, row=nlwsn, value=row[0].value)
                wsn.cell(column=3, row=nlwsn, value=row[1].value)
                wsn.cell(column=4, row=nlwsn, value=row[2].value)
                wsn.cell(column=10, row=nlwsn, value=row[9].value)
                print(row[0].value + " " + row[6].value)

    except (TypeError, AttributeError):
        continue

bold12Font = Font(size=12, bold=True)
border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

row1 = "A1:G1"
for sheet in wb:
    for c in sheet[row1]:
        for a in c:
            a.font = bold12Font
rows = "A1:K36"
for sheet in wb:
    for row in sheet[rows]:
        for cell in row:
            cell.border = border
row2 = "A2:A37"
for sheet in wb:
    for row in sheet[row2]:
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

for sheet in wb:
    sheet['A2'] = 1
    sheet['A3'] = 2
    sheet['A4'] = 3
    sheet['A5'] = 4
    sheet['A6'] = 5
    sheet['A7'] = 6
    sheet['A8'] = 7
    sheet['A9'] = 8
    sheet['A10'] = 9
    sheet['A11'] = 10
    sheet['A12'] = 11
    sheet['A13'] = 12
    sheet['A14'] = 13
    sheet['A15'] = 14
    sheet['A16'] = 15
    sheet['A17'] = 16
    sheet['A18'] = 17
    sheet['A19'] = 18
    sheet['A20'] = 19
    sheet['A21'] = 20
    sheet['A22'] = 21
    sheet['A23'] = 22
    sheet['A24'] = 23
    sheet['A25'] = 24
    sheet['A26'] = 25
    sheet['A27'] = 26
    sheet['A28'] = 27
    sheet['A29'] = 28
    sheet['A30'] = 29
    sheet['A31'] = 30
    sheet['A32'] = 31
    sheet['A33'] = 32
    sheet['A34'] = 33
    sheet['A35'] = 34
    sheet['A36'] = 35
    sheet['A37'] = 36

    
ws.oddHeader.left.text = 'QC_SlideMaking_DSP' + str(date.today() + timedelta(days=1)) 
ws.oddHeader.left.size = 12
ws.oddHeader.left.font = "Tahoma,Bold"
#ws.oddHeader.center.color = "CC3366"

#ws.oddHeader.right.text = "Tech Loading: ________________ \r Tech Checking: ________________"
#ws.oddHeader.right.size = 12
#ws.oddHeader.right.font = "Tahoma,Bold"

#ws.oddHeader.center.text = "Date: ________________"
#ws.oddHeader.center.size = 12
#ws.oddHeader.center.font = "Tahoma,Bold"

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

wb.save('QC_SlideMaking_DSP Week of' + '_' + str(last_monday) + '.xlsx')
wb.close()

'''row1 = ws["A1:G1"]
for c in row1:
    for a in c:
        a.font = bold12Font     
rows = ws["A1:K36"]
for row in rows:
    for cell in row:
        cell.border = border
        #cell.font = bold12Font
        #cell.alignment = Alignment(horizontal="center")
        #cell.font = "Tahoma,Bold"
row2 = ws["A2:A31"]
for row in row2:
    for cell in row:
        cell.alignment = Alignment(horizontal="center")'''


'''ws['A2'] = 1
ws['A3'] = 2
ws['A4'] = 3
ws['A5'] = 4
ws['A6'] = 5
ws['A7'] = 6
ws['A8'] = 7
ws['A9'] = 8
ws['A10'] = 9
ws['A11'] = 10
ws['A12'] = 11
ws['A13'] = 12
ws['A14'] = 13
ws['A15'] = 14
ws['A16'] = 15
ws['A17'] = 16
ws['A18'] = 17
ws['A19'] = 18
ws['A20'] = 19
ws['A21'] = 20
ws['A22'] = 21
ws['A23'] = 22
ws['A24'] = 23
ws['A25'] = 24
ws['A26'] = 25
ws['A27'] = 26
ws['A28'] = 27
ws['A29'] = 28
ws['A30'] = 29
ws['A31'] = 30'''

#x = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
#colnum = ws["K2:K30"]
#for row in colnum:
 #   for col in row:
  #      ws.cell(column=11, row=ws.min_row+1, value=x)

#row2 = ws["E2:K36"]
#for r in row2:
 #   for d in r:
  #      if d.value == None:
   #         ws.delete_cols(index)
        
#for row in ws.rows:
 #   cell = row[:8][1]
  #  cell1 = row[:8][2]
   # cell2 = row[:8][3]
   # cell3 = row[:8][4]
    #cell4 = row[8:][0]
#    cell.alignment = Alignment(horizontal="center")
 #   cell1.alignment = Alignment(horizontal="center")
  #  cell2.alignment = Alignment(horizontal="center")
   # cell3.alignment = Alignment(horizontal="center")
   # cell4.alignment = Alignment(horizontal="center")

#wb.remove(wb['Sheet'])
#ws.delete_rows(3,4)


